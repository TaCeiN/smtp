import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
import xlrd


# Импорт БД
def excel():
    print("Выберите тип файла Excel")
    while True:
        excel_type = input("1 - .xlsx \t 2 - .xls\n")
        if excel_type in ['1', '2']:
            break

        print("Пожалуйста, выберите 1 или 2.")

    while True:
        DB_path = input("Название файла Excel: ")
        try:

            if excel_type == '1':
                workbook = openpyxl.load_workbook(f"{DB_path}.xlsx")
            else:
                workbook = openpyxl.load_workbook(f"{DB_path}.xls")

            return workbook.active
        except FileNotFoundError:
            print("Файл не найден. Пожалуйста, попробуйте снова.")


def PDF():
    while True:
        File_path = input("Название файла PDF: ")
        try:
            open(f"{File_path}.pdf")
            return File_path
        except FileNotFoundError:
            print("Файл не найден. Пожалуйста, попробуйте снова.")


active_workbook = excel()
File_path = PDF()

# Логин/Пароль/Файл
gmail_user = '[EMAIL]'
gmail_app_password = '[EMAIL_PASSWORD]'
filename = f"{File_path}.pdf"

# Получатели email и ФИО
sender_to = []
name = []

def email_name_xlsx(active_workbook):
    for i in range(1, active_workbook.max_row + 1):
        sender_to.append(active_workbook.cell(i, 1).value)
    for j in range(1, active_workbook.max_row + 1):
        name.append(active_workbook.cell(j, 2).value)

email_name_xlsx(active_workbook)

# Письмо
for i, name in zip(sender_to, name):
    # Основная часть и Заголовок
    msg = MIMEMultipart()
    msg['From'] = '[YOUR_SENDER_EMAIL]'
    msg['To'] = i
    msg['Subject'] = "[YOUR_TOPIC]"
    # Тело Письма
    body = f"""
    <p> Уважаемый {name} !<br>
    [EMAIL_BODY].<br>
    </p>
    """

    # Подпись в письме
    signature = """
    <p>С уважением,<br>
    [SIGNATURE]
    </p>
    """

    # Прикрепление файла к письму
    with open(filename, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        msg.attach(part)
    # Приклепление Тела и Подписи
    msg.attach(MIMEText(body, 'html', 'utf-8'))
    msg.attach(MIMEText(signature, 'html', 'utf-8'))

    # Попытка отправить письмо с обработкой исключений
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(gmail_user, gmail_app_password)
            server.sendmail(gmail_user, i, msg.as_string())
        print('Email sent!')
    except Exception as exception:
        print("Error: %s!\n\n" % exception)

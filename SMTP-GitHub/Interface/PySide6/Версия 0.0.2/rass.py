import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from fileinput import filename
from main0_3 import *
import openpyxl
import xlrd


def read_value_from_txt(txtname):
    values = {}
    with open(txtname, 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()
            line = line.split('#')[0].strip()
            if line:
                key, value = line.split(":")
                value = value.strip("'")
                if value.isdigit():
                    value = int(value)
                else:
                    try:
                        value = float(value)
                    except ValueError:
                        pass
                values[key] = value
    return values

from main0_3 import *
txtname = 'config.txt'
values = read_value_from_txt(txtname)
sender_to = []
name_list = []
user = values['user']
app_password = values['app_password']
topic = values['topic']
first_line = values['first_line']
mail = values['mail']
dear = values['apply']
port = values['port']


# Выбор файла(не более 1) и выбор типа обработки(в зависимости от формата)
def excel_format(db_file_names):
    if len(db_file_names) == 1:
        for file in db_file_names:
            file = file.split('.')[1].strip()
            if file == "xlsx":
                sender_to.clear(),name_list.clear()
                workbook = openpyxl.load_workbook(f"{db_file_names[0]}")
                active_workbook = workbook.active
                email_info_xlsx(active_workbook)
                print(f"Выбран файл{db_file_names[0]}")
                print(name_list,sender_to)
                return workbook.active,True
            if file == "xls":
                sender_to.clear(),name_list.clear()
                workbook = xlrd.open_workbook(f"{db_file_names[0]}")
                sheet = workbook.sheet_by_index(0)
                email_info_xls(sheet)
                print(f"Выбран файл{db_file_names[0]}")
                print(name_list, sender_to)
                return sheet,True
    else:
        print("Выберано 2 и более файлов")
        return False

# Обработка XlSX
def email_info_xlsx(active_workbook):
    for i in range(1, active_workbook.max_row + 1):
        sender_to.append(active_workbook.cell(i, 1).value)
        name_list.append(active_workbook.cell(i, 2).value)
    return sender_to, name_list


# Обработка XLS
def email_info_xls(sheet):
    for i in range(sheet.nrows):
        sender_to.append(sheet.cell_value(i, 0))
        name_list.append(sheet.cell_value(i, 1))
    return sender_to, name_list




# Письмо
def send_email():
    for i, name in zip(sender_to, name_list):
        # Основная часть и Заголовок
        msg = MIMEMultipart()
        msg['From'] = user
        msg['To'] = i
        msg['Subject'] = f'{topic}'
        # Тело Письма
        body = f"""
        <p> {dear} {name}<br>
        {first_line}.<br>
        <img src="https://mc.yandex.ru/pixel/4758594046811402458?rnd=%aw_random%"> <br>
        </p>
        """

        # Подпись в письме
        signature = """
        <p>С уважением,<br>
        [SIGNATURE]
        </p>
        """

        # Прикрепление файла к письму
        with open(filename, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)
        # Приклепление Тела и Подписи
        msg.attach(MIMEText(body, 'html', 'utf-8'))
        msg.attach(MIMEText(signature, 'html', 'utf-8'))

        # Попытка отправить письмо с обработкой исключений
        try:
            with smtplib.SMTP_SSL(f'smtp.{mail}.{port}', 465) as server:
                server.login(user, app_password)
                server.sendmail(user, i, msg.as_string())
            print('Письмо отправлено')
        except Exception as exception:
            print("Error: %s!\n\n" % exception)



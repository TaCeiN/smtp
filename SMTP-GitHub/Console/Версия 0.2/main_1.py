import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import openpyxl

# Импорт БД
workbook = openpyxl.load_workbook("BD.xlsx")
active_workbook = workbook.active

# Логин/Пароль/Файл
gmail_user = '[EMAIL]'
gmail_app_password = '[EMAIL_PASSWORD]'
filename = '[ATTACH_FILE]'

# Получатели email и ФИО
sender_to = []
name = []
for i in range(1, active_workbook.max_row + 1):
    sender_to.append(active_workbook.cell(i, 1).value)
for j in range(1, active_workbook.max_row + 1):
    name.append(active_workbook.cell(j, 2).value)

# Письмо
for i, j in zip(sender_to,name):
    # Основная часть и Заголовок
    msg = MIMEMultipart()
    msg['From'] = '[YOUR_SENDER_EMAIL]'
    msg['To'] = i
    msg['Subject'] = "[YOUR_TOPIC]"
    # Тело Письма
    body = f"""
    <p> DEAR {j} !<br>
    [EMAIL_BODY].<br>
    </p>
    """

    # Подпись в письме
    signature = """
    <p>С уважением,<br>
    [SIGNATURE]
    </p>
    """

    # Прикрепление файла к письму
    with open(filename, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        msg.attach(part)
    # Приклепление Тела и Подписи
    msg.attach(MIMEText(body, 'html', 'utf-8'))
    msg.attach(MIMEText(signature, 'html', 'utf-8'))

    # Попытка отправить письмо с обработкой исключений
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(gmail_user, gmail_app_password)
            server.sendmail(gmail_user, i, msg.as_string())
        print('Email sent!')
    except Exception as exception:
        print("Error: %s!\n\n" % exception)

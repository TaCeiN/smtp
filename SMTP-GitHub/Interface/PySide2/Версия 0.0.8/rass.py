import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import Header
import os
from inspect import signature

import main0_3
import openpyxl
import xlrd



def read_value_from_txt(txtname):
    values = {}
    with open(txtname, 'r', encoding='utf-8') as file:
        current_key = None
        current_value = []

        for line in file:
            line = line.strip()
            line = line.split('#')[0].strip()  # Убираем комментарии

            if not line:
                continue  # Пропускаем пустые строки

            # Проверяем, начинается ли строка с нового ключа
            if ':' in line:
                if current_key is not None:
                    # Сохраняем предыдущее значение
                    values[current_key] = ' '.join(current_value).strip().rstrip("'")  # Убираем лишнюю кавычку

                # Извлекаем новый ключ и значение
                key, value = line.split(":", 1)
                current_key = key.strip()
                current_value = [value.strip().strip("'")]  # Начинаем новую запись

            else:
                # Если строка не содержит ключа, добавляем к текущему значению
                current_value.append(line.strip())

        # Сохраняем последнее значение после завершения цикла
        if current_key is not None:
            values[current_key] = ' '.join(current_value).strip().rstrip("'")  # Убираем лишнюю кавычку в конце

        # Обработка значений
        for key, value in values.items():
            if value.isdigit():
                values[key] = int(value)
            else:
                try:
                    values[key] = float(value)
                except ValueError:
                    pass  # Если не число, оставляем как есть

            # Проверяем на наличие URL (или других специальных строк)
            if value.startswith("http://") or value.startswith("https://"):
                pass  # Оставляем значение без изменений
            elif "%aw_random%" in value:
                pass  # Оставляем значение без изменений

    return values


def values_from_txt(txtname):
    txtname = 'config.txt'
    values = read_value_from_txt(txtname)
    user = values['user']
    app_password = values['app_password']
    text = values['body']
    signature = values['signature']
    topic = values['topic']
    pixel = values['pixel']
    return sender_to,name_list,user,app_password,text,signature,topic,pixel

file_count = 0
sender_to = []
name_list = []

################################################    select_db_file

# Выбор файла(не более 1) и выбор типа обработки(в зависимости от формата)
def excel_format(db_file_names,db_filenames):
    if len(db_file_names) == 1:
        for file in db_file_names:
            file = file.split('.')[1].strip()
            if file == "xlsx":
                sender_to.clear(),name_list.clear()
                workbook = openpyxl.load_workbook(db_filenames)
                active_workbook = workbook.active
                email_info_xlsx(active_workbook)
                print(f"Выбран файл{db_file_names[0]}")
                print(name_list,sender_to)
                return workbook.active,True
            if file == "xls":
                sender_to.clear(),name_list.clear()
                workbook = xlrd.open_workbook(db_filenames)
                sheet = workbook.sheet_by_index(0)
                email_info_xls(sheet)
                print(f"Выбран файл{db_file_names[0]}")
                print(name_list, sender_to)
                return sheet,True
    else:
        print("Выберано 2 и более файлов")
        return False

# Обработка XlSX
def email_info_xlsx(active_workbook):
    for i in range(1, active_workbook.max_row + 1):
        sender_to.append(active_workbook.cell(i, 1).value)
        name_list.append(active_workbook.cell(i, 2).value)
    return sender_to, name_list

# Обработка XLS
def email_info_xls(sheet):
    for i in range(sheet.nrows):
        sender_to.append(sheet.cell_value(i, 0))
        name_list.append(sheet.cell_value(i, 1))
    return sender_to, name_list

################################################    select_db_file


def attach_file_save(filenames):
    global file_count
    file_count = len(filenames)
    filename_t = filenames
    print(filename_t)
    return filename_t


def send_email(i, email_number, filename_t):
    global file_count
    sender_to,name_list,user,app_password,text,signature,topic,pixel = values_from_txt("config.txt")
    # Основная часть и Заголовок
    msg = MIMEMultipart()
    msg['From'] = user
    msg['To'] = i
    msg['Subject'] = f'{topic}'
    # Тело Письма
    body = f"""
    {text.replace('{name_list}', name_list[email_number])}
    """

    # Подпись в письме
    signatur = f"""
        {signature}
    """
    for attach in filename_t:
        # Прикрепление файла к письму
        with open(attach, 'rb') as attachment:
            part = MIMEApplication(attachment.read(), Name=os.path.basename(attach))
            part['Content-Disposition'] = f'attachment; filename="{Header(os.path.basename(attach), "utf-8").encode()}"'
            msg.attach(part)
    # Приклепление Тела и Подписи
    msg.attach(MIMEText(body, 'html', 'utf-8'))
    msg.attach(MIMEText(signatur, 'html', 'utf-8'))

    # Попытка отправить письмо с обработкой исключений
    try:
        with smtplib.SMTP_SSL(f'smtp.gmail.com', 465) as server:
            server.login(user, app_password)
            server.sendmail(user, i, msg.as_string())
        print('Письмо отправлено')
    except Exception as exception:
        print("Error: %s!\n\n" % exception)



def sendd(update_progress,filename_t):
    email_itera = int(100/len(sender_to))
    email_number = 0
    maximum = len(sender_to)*email_itera
    for i, name in zip(sender_to, name_list):
        send_email(i,email_number, filename_t)
        email_number+=1
        email_progress_value = email_itera*email_number
        update_progress(email_progress_value,maximum)

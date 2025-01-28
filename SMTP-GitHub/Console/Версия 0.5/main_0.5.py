import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
import xlrd

# Email,Обращение,Данные аккаунта
sender_to = []
name = []
gmail_user = '[EMAIL]'
gmail_app_password = '[EMAIL_PASSWORD]'

# Проверка формата файла Excel(1 или 2)
def excel_format():
    print("Выберите тип файла Excel")
    while True:
        excel_type = input("1 - .xlsx \t 2 - .xls\n")
        if excel_type in ['1', '2']:
            return excel_type

        print("Пожалуйста, выберите 1 или 2.")

# Открытие файла Execl в зависимости от формата
def excel(excel_type):
    while True:
        DB_path = input("Название файла Excel: ")
        try:

            if excel_type == '1':
                workbook = openpyxl.load_workbook(f"{DB_path}.xlsx")
                return workbook.active
            else:
                workbook = xlrd.open_workbook(f"{DB_path}.xls")
                sheet = workbook.sheet_by_index(0)
                return sheet

        except FileNotFoundError:
            print("Файл не найден. Пожалуйста, попробуйте снова.")

#Открытие PDF-файла
def PDF():
    while True:
        File_path = input("Название файла PDF: ")
        try:
            open(f"{File_path}.pdf")
            return File_path
        except FileNotFoundError:
            print("Файл не найден. Пожалуйста, попробуйте снова.")


# Обработка XlSX
def email_info_xlsx(active_workbook):
    for i in range(1, active_workbook.max_row + 1):
        sender_to.append(active_workbook.cell(i, 1).value)
    for j in range(1, active_workbook.max_row + 1):
        name.append(active_workbook.cell(j, 2).value)
    return sender_to, name


# Обработка XLS
def email_info_xls(sheet):
    for i in range(sheet.nrows):
        sender_to.append(sheet.cell_value(i, 0))
    for j in range(sheet.nrows):
        name.append(sheet.cell_value(j, 1))
    return sender_to, name

# Получение формата таблицы и обработка в зависимости от формата
excel_type = excel_format()
if excel_type == '1':
    active_workbook = excel(excel_type)
    email_info_xlsx(active_workbook)
if excel_type == '2':
    sheet = excel(excel_type)
    email_info_xls(sheet)

File_path = PDF()
filename = f"{File_path}.pdf"

# Письмо
for i, name in zip(sender_to, name):
    # Основная часть и Заголовок
    msg = MIMEMultipart()
    msg['From'] = gmail_user
    msg['To'] = i
    msg['Subject'] = "[YOUR_TOPIC]"
    # Тело Письма
    body = f"""
    <p> Уважаемый {name} !<br>
    [EMAIL_BODY].<br>
    </p>
    """

    # Подпись в письме
    signature = """
    <p>С уважением,<br>
    [SIGNATURE]
    </p>
    """

    # Прикрепление файла к письму
    with open(filename, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        msg.attach(part)
    # Приклепление Тела и Подписи
    msg.attach(MIMEText(body, 'html', 'utf-8'))
    msg.attach(MIMEText(signature, 'html', 'utf-8'))

    # Попытка отправить письмо с обработкой исключений
    try:
        with smtplib.SMTP_SSL('smtp.mail.ru', 465) as server:
            server.login(gmail_user, gmail_app_password)
            server.sendmail(gmail_user, i, msg.as_string())
        print('Email sent!')
    except Exception as exception:
        print("Error: %s!\n\n" % exception)

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import Header
import os
import main0_3
import openpyxl
import xlrd


def read_value_from_txt(txtname):
    values = {}
    with open(txtname, 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()
            line = line.split('#')[0].strip()  # Убираем комментарии
            if line:
                key, value = line.split(":", 1)  # Разделяем только на первое вхождение ':'
                value = value.strip("'").strip()  # Убираем одинарные кавычки и пробелы

                # Проверяем, является ли значение числом
                if value.isdigit():
                    value = int(value)
                else:
                    try:
                        value = float(value)
                    except ValueError:
                        pass  # Если не число, оставляем как есть

                # Проверяем на наличие URL (или других специальных строк)
                if value.startswith("http://") or value.startswith("https://"):
                    pass  # Оставляем значение без изменений
                elif "%aw_random%" in value:
                    pass  # Оставляем значение без изменений

                values[key] = value
    return values


txtname = 'config.txt'
values = read_value_from_txt(txtname)
sender_to = []
name_list = []
user = values['user']
app_password = values['app_password']
topic = values['topic']
first_line = values['first_line']
mail = values['mail']
dear = values['apply']
port = values['port']
pixel = values['pixel']
file_count = 0


################################################    select_db_file

# Выбор файла(не более 1) и выбор типа обработки(в зависимости от формата)
def excel_format(db_file_names,db_filenames):
    if len(db_file_names) == 1:
        for file in db_file_names:
            file = file.split('.')[1].strip()
            if file == "xlsx":
                sender_to.clear(),name_list.clear()
                workbook = openpyxl.load_workbook(db_filenames)
                active_workbook = workbook.active
                email_info_xlsx(active_workbook)
                print(f"Выбран файл{db_file_names[0]}")
                print(name_list,sender_to)
                return workbook.active,True
            if file == "xls":
                sender_to.clear(),name_list.clear()
                workbook = xlrd.open_workbook(db_filenames)
                sheet = workbook.sheet_by_index(0)
                email_info_xls(sheet)
                print(f"Выбран файл{db_file_names[0]}")
                print(name_list, sender_to)
                return sheet,True
    else:
        print("Выберано 2 и более файлов")
        return False

# Обработка XlSX
def email_info_xlsx(active_workbook):
    for i in range(1, active_workbook.max_row + 1):
        sender_to.append(active_workbook.cell(i, 1).value)
        name_list.append(active_workbook.cell(i, 2).value)
    return sender_to, name_list

# Обработка XLS
def email_info_xls(sheet):
    for i in range(sheet.nrows):
        sender_to.append(sheet.cell_value(i, 0))
        name_list.append(sheet.cell_value(i, 1))
    return sender_to, name_list

################################################    select_db_file


def attach_file_save(filenames):
    global file_count
    file_count = len(filenames)
    filename_t = filenames
    print(filename_t)
    return filename_t


def send_email(i, name, filename_t):
    global file_count
    # Основная часть и Заголовок
    msg = MIMEMultipart()
    msg['From'] = user
    msg['To'] = i
    msg['Subject'] = f'{topic}'
    # Тело Письма
    body = f"""
    <p> {dear} {name}<br>
    {first_line}.<br>
    <img src="{pixel}"> <br>
    </p>
    """

    # Подпись в письме
    signature = """
    <p>С уважением,<br>
    [SIGNATURE]
    </p>
    """
    for attach in filename_t:
        # Прикрепление файла к письму
        with open(attach, 'rb') as attachment:
            part = MIMEApplication(attachment.read(), Name=os.path.basename(attach))
            part['Content-Disposition'] = f'attachment; filename="{Header(os.path.basename(attach), "utf-8").encode()}"'
            msg.attach(part)
    # Приклепление Тела и Подписи
    msg.attach(MIMEText(body, 'html', 'utf-8'))
    msg.attach(MIMEText(signature, 'html', 'utf-8'))

    # Попытка отправить письмо с обработкой исключений
    try:
        with smtplib.SMTP_SSL(f'smtp.{mail}.{port}', 465) as server:
            server.login(user, app_password)
            server.sendmail(user, i, msg.as_string())
        print('Письмо отправлено')
    except Exception as exception:
        print("Error: %s!\n\n" % exception)



def sendd(update_progress,filename_t):
    email_itera = int(100/len(sender_to))
    email_number = 0
    maximum = len(sender_to)*email_itera
    for i, name in zip(sender_to, name_list):
        send_email(i, name, filename_t)
        email_number+=1
        email_progress_value = email_itera*email_number
        update_progress(email_progress_value,maximum)
